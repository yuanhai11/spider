import openpyxl
import os,time
from datetime import datetime

from xlrd import open_workbook
import pywintypes
####  pyinstaller 打包时出现问题，下载pywin32types即可，pip install pypiwin32


def parse():
    # 获取用户目录
    path = os.listdir('.')
    path = [i for i in path if "xls" in i]
    print(path)
    for p in path:

        workbook = open_workbook(p)  # 打开excel文件
        sheet2 = workbook.sheets()
        print(sheet2)
        # exit()
        num = 1
        listss = []
        sheet2 = workbook.sheet_by_index(0)
        print(sheet2.nrows)
        # title = str(sheet2.cell(0, 0).value).strip().replace(' ','')
        # print(title)

        for i in range(2, sheet2.nrows):

            phone = sheet2.cell(i, 9).value
            more_phone = sheet2.cell(i, 10).value
            ss = []
            if more_phone == '-':
                ss.append(phone)
            else:
                ss += more_phone.split('；')
                ss.append(phone)
            ss = list(set(ss))
            for jjj in ss:
                if '-' in jjj or jjj.startswith('400'):
                    continue
                else:
                    listss.append(jjj)

        listss = list(set(listss))
        for jj in listss:
            with open('data.txt','a',encoding='utf-8') as fp:
                fp.write(jj+'\n')

def parse2():
    import pandas as pd
    df = pd.read_excel('查企业_高级搜索_企查查(62952169).xlsx')  # 这个会直接默认读取到这个Excel的第一个表单
    print(df)

if __name__ == '__main__':
    # parse()
    parse2()