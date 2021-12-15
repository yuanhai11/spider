
import openpyxl
from xlrd import open_workbook

def parse():

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    workbook = open_workbook('compass.xls')  # 打开excel文件
    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1

    for index,sheet in enumerate(sheet2):

        sheet2 = workbook.sheet_by_index(index)
        print(sheet2.nrows)
        for i in range(1, sheet2.nrows):

            title_0 = sheet2.cell(i, 0).value
            title_1 = sheet2.cell(i, 1).value
            title_2 = sheet2.cell(i, 4).value
            title_3 = sheet2.cell(i, 10).value
            title_4 = sheet2.cell(i, 12).value
            title_5 = sheet2.cell(i, 13).value

            sum = ['1',title_0,
                   title_1,
                   title_2,
                   title_3,
                   title_4,
                   title_5,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1

    outwb.save('compass_done.xlsx')
if __name__ == '__main__':
    parse()