import time, os
import logging.handlers
import openpyxl
# import pywintypes

####  pyinstaller 打包时出现问题，下载pywin32types即可，pip install pypiwin32

def parse():
    # 获取用户目录
    LOCAL_FILE_PATH = input('输入文件')

    new_files = LOCAL_FILE_PATH.split('.')[0]

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    workbook = open_workbook(LOCAL_FILE_PATH)  # 打开excel文件
    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1
    for index,sss in enumerate(sheet2):

        sheet2 = workbook.sheet_by_index(index)
        print(sheet2.nrows)
        for i in range(1, sheet2.nrows):
            title_01 = sheet2.cell(i, 0).value
            title_0 = sheet2.cell(i, 1).value
            title_1 = sheet2.cell(i, 2).value
            title_2 = sheet2.cell(i, 3).value
            title_3 = sheet2.cell(i, 4).value
            title_4 = sheet2.cell(i, 5).value
            title_5 = sheet2.cell(i, 6).value
            title_6 = sheet2.cell(i, 7).value
            title_7 = sheet2.cell(i, 8).value
            title_8 = sheet2.cell(i, 9).value

            title_9 = sheet2.cell(i, 10).value

            title_10 = sheet2.cell(i, 11).value
            title_11 = sheet2.cell(i, 12).value
            title_12 = sheet2.cell(i, 13).value
            title_13 = sheet2.cell(i, 14).value
            title_14 = sheet2.cell(i, 15).value
            title_15 = sheet2.cell(i, 16).value

            more_phone = title_9
            ss = []
            if more_phone == '-':
                continue
            else:
                ss += more_phone.split(',')

            ss = list(set(ss))
            ss = [d for d in ss if '暂不公示' not in d and '-' not in d and d[0]!='0']
            # continue
            # regis_date = sheet2.cell(i, 3).value
            # c = datetime(*xldate_as_tuple(regis_date, 0))
            # c = c.strftime('%Y-%m-%d')
            # print(title_20)
            for jj in ss:

                sum = ['1', title_01,
                       title_0,
                       title_1,
                       title_2,
                       title_3,
                       title_4,
                       title_5,
                       title_6,
                       title_7,
                       title_8,
                       jj,
                       title_10,
                       title_11,
                       title_12,
                       title_13,
                       title_14,
                       title_15,
                       ]
                for index, s in enumerate(sum):
                    outws.cell(num, index + 1).value = s  #
                num += 1

    outwb.save('{}_done1.xlsx'.format(new_files))

if __name__ == '__main__':
    parse()