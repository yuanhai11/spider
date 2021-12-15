import time, os
import logging.handlers
import openpyxl
# import pywintypes

####  pyinstaller 打包时出现问题，下载pywin32types即可，pip install pypiwin32

def parse():
    # 获取用户目录
    LOCAL_FILE_PATH = input('输入文件')

    new_files = LOCAL_FILE_PATH.split('.')[0]

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    workbook = open_workbook(LOCAL_FILE_PATH)  # 打开excel文件
    sheet2 = workbook.sheets()
    print(sheet2)
    # exit()
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)
    title = str(sheet2.cell(0, 0).value).strip().replace(' ','')
    print(title)
    if '企查查' in title:
        for i in range(2, sheet2.nrows):

            title_01 = sheet2.cell(i, 0).value
            title_0 = sheet2.cell(i, 1).value
            title_1 = sheet2.cell(i, 2).value
            title_2 = sheet2.cell(i, 3).value
            title_3 = sheet2.cell(i, 4).value
            title_4 = sheet2.cell(i, 5).value
            title_5 = sheet2.cell(i, 6).value
            title_6 = sheet2.cell(i, 7).value
            title_7 = sheet2.cell(i, 8).value

            title_8 = sheet2.cell(i, 9).value
            title_9 = sheet2.cell(i, 10).value

            title_10 = sheet2.cell(i, 11).value
            phone = title_8
            more_phone = title_9
            ss = []
            if more_phone == '-':
                ss.append(phone)
            else:
                ss += more_phone.split('；')
                ss.append(phone)
            ss = list(set(ss))
            # continue
            # regis_date = sheet2.cell(i, 3).value
            # c = datetime(*xldate_as_tuple(regis_date, 0))
            # c = c.strftime('%Y-%m-%d')
            # print(title_20)
            for jj in ss:

                sum = ['1', title_01,
                       title_0,
                       title_1,
                       title_2,
                       title_3,
                       title_4,
                       title_5,
                       title_6,
                       title_7,
                       jj,
                       title_9,
                       title_10,
                       title_11,
                       title_12,
                       title_13,
                       title_14,
                       title_15,
                       title_16,
                       title_17,
                       title_18,
                       title_19,
                       title_20,
                       title_21,
                       title_22,
                       title_23,
                       title_24,
                       ]
                for index, s in enumerate(sum):
                    outws.cell(num, index + 1).value = s  #
                num += 1
    else:
        for i in range(3, sheet2.nrows):

            title_01 = sheet2.cell(i, 0).value
            title_0 = sheet2.cell(i, 1).value
            title_1 = sheet2.cell(i, 2).value
            title_2 = sheet2.cell(i, 3).value
            title_3 = sheet2.cell(i, 4).value
            title_4 = sheet2.cell(i, 5).value
            title_5 = sheet2.cell(i, 6).value
            title_6 = sheet2.cell(i, 7).value
            title_7 = sheet2.cell(i, 8).value

            title_8 = sheet2.cell(i, 9).value
            title_9 = sheet2.cell(i, 10).value
            # print(type(title_8),title_8,type(title_9),title_9)
            # continue
            title_10 = sheet2.cell(i, 11).value
            title_11 = sheet2.cell(i, 12).value
            title_12 = sheet2.cell(i, 13).value
            title_13 = sheet2.cell(i, 14).value
            title_14 = sheet2.cell(i, 15).value
            title_15 = sheet2.cell(i, 16).value
            title_16 = sheet2.cell(i, 17).value
            title_17 = sheet2.cell(i, 18).value
            title_18 = sheet2.cell(i, 19).value
            title_19 = sheet2.cell(i, 20).value
            title_20 = sheet2.cell(i, 21).value

            phone = sheet2.cell(i, 22).value.split(';')
            invalid_phone = sheet2.cell(i, 23).value
            more_phone = sheet2.cell(i, 24).value
            ss = []
            if more_phone == '-':
                ss += phone
            else:
                ss += more_phone.split(';')
                ss += phone

            title_24 = sheet2.cell(i, 25).value
            title_25 = sheet2.cell(i, 26).value
            title_26 = sheet2.cell(i, 27).value

            # continue
            # regis_date = sheet2.cell(i, 3).value
            # c = datetime(*xldate_as_tuple(regis_date, 0))
            # c = c.strftime('%Y-%m-%d')
            # print(title_20)
            for jj in ss:

                sum = ['1', title_01,
                       title_0,
                       title_1,
                       title_2,
                       title_3,
                       title_4,
                       title_5,
                       title_6,
                       title_7,
                       title_8,
                       title_9,
                       title_10,
                       title_11,
                       title_12,
                       title_13,
                       title_14,
                       title_15,
                       title_16,
                       title_17,
                       title_18,
                       title_19,
                       title_20,
                       jj,
                       invalid_phone,
                       more_phone,
                       title_24,
                       title_25,
                       title_26,
                       ]
                for index, s in enumerate(sum):
                    outws.cell(num, index + 1).value = s  #
                num += 1

    outwb.save('{}_done.xlsx'.format(new_files))

if __name__ == '__main__':
    parse()