import json

import openpyxl

def parse():



    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    workbook = open_workbook("b_done.xlsx")  # 打开excel文件
    workbook1 = open_workbook("a.xlsx")  # 打开excel文件



    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)

    num1 = 1
    sheet3 = workbook1.sheet_by_index(0)
    print(sheet3.nrows)

    for i in range(1, sheet2.nrows):
            company_name = sheet2.cell(i, 0).value
            phone = sheet2.cell(i, 1).value
            for j in range(1,sheet3.nrows):
                company_name1 = sheet3.cell(j, 0).value
                phone1 = sheet3.cell(j,1).value

                if company_name==company_name1:
                    pass
                    phone = phone + ',' + phone1
                else:
                    continue


            ss = ','.join(list(set(phone.split(','))))
            # for jj in ss:

            sum = ['1', company_name,
                   ss,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1

    outwb.save('{}_done1.xlsx'.format('b'))


def parse2():
    from xlrd import open_workbook
    workbook = open_workbook("b_done.xlsx")  # 打开excel文件
    workbook1 = open_workbook("a.xlsx")  # 打开excel文件

    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)

    num1 = 1
    sheet3 = workbook1.sheet_by_index(0)
    print(sheet3.nrows)
    a = []
    b = []
    c = []
    for i in range(0,sheet2.nrows):
        a.append(sheet2.cell(i, 0).value)
    print(a)
    for j in range(1,sheet3.nrows):
        jj = sheet3.cell(j, 1).value
        if jj not in a:
            b.append(jj)
        else:
            c.append(jj)
    print(b)
    print(len(b))
    print(c)


def parse3():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    workbook = open_workbook("b_done1.xlsx")  # 打开excel文件

    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)
    for i in range(1,sheet2.nrows):
        company_name = sheet2.cell(i, 0).value
        phone = sheet2.cell(i, 1).value
        try:
            if ',' in phone:
                phone = phone.split(',')
                # for jj in ss:
                for h in phone:
                    sum = ['1', company_name,
                           h,
                           ]
                    for index, s in enumerate(sum):
                        outws.cell(num, index + 1).value = s  #
                    num += 1
            else:
                sum = ['1', company_name,
                       phone,
                       ]
                for index, s in enumerate(sum):
                    outws.cell(num, index + 1).value = s  #
                num += 1

        except Exception:
            sum = ['1', company_name,
                   phone,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1



    outwb.save('{}_done11.xlsx'.format('b'))

def parse4():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    workbook = open_workbook("b_done11.xlsx")  # 打开excel文件

    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)
    a=[]
    for i in range(1,sheet2.nrows):
        company_name = sheet2.cell(i, 0).value
        phone = sheet2.cell(i, 1).value
        a.append(phone)
    a = list(set(a))

    b = []
    for jj in a:
        k = 0
        for i in range(1, sheet2.nrows):
            phone = sheet2.cell(i, 1).value
            if phone == jj:
                k+=1
        if k>=10:
            b.append(jj)


    for ii in range(1,sheet2.nrows):
        company_name = sheet2.cell(ii, 0).value
        phone = sheet2.cell(ii, 1).value
        if phone in b:
            print(phone,"重复次数达到10")
            continue

        sum = ['1', company_name,
               phone,
               ]
        for index, s in enumerate(sum):
            outws.cell(num, index + 1).value = s  #
        num += 1

    outwb.save('{}_done1111.xlsx'.format('b'))


if __name__ == '__main__':
    # parse()
    # parse2()
    # parse3()
    parse4()