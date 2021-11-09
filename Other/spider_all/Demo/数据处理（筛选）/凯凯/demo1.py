import openpyxl

####  pyinstaller 打包时出现问题，下载pywin32types即可，pip install pypiwin32

def parse():

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    from xlrd import open_workbook
    workbook = open_workbook("新建 XLSX 工作表.xlsx")  # 打开excel文件
    sheet2 = workbook.sheets()
    print(sheet2)
    # exit()
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)
    sum = []
    sum1 = []
    for i in range(0,sheet2.nrows):
        title_01 = sheet2.cell(i, 1).value
        title_0 = sheet2.cell(i, 3).value
        sum.append(title_01)
        sum1.append(title_0)

    sum = [i for i in sum if i != '']
    print(sum)
    # sum = [i[0:14] for i in sum if i != '']
    # print(sum)
    print(sum1)
    s = []
    s1 = []
    for j in sum:
        if j[0:14] not in sum1:
            s.append(j)

        else:
            s1.append(j)


    print("不在",s)
    print("在",s1)



if __name__ == '__main__':

    parse()