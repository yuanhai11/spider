import time, os
import logging.handlers
import openpyxl
# import pywintypes

####  pyinstaller 打包时出现问题，下载pywin32types即可，pip install pypiwin32

import time
import re
import json
import requests, pymysql
from sqlalchemy import Column, String, create_engine, Integer
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base

# 创建对象的基类:
Base = declarative_base()


# 定义User对象:

class Medicine(Base):
    # 表的名字:
    __tablename__ = 'icp_leads_copy1'
    id = Column(Integer(), primary_key=True, autoincrement=True)
    state_code = Column(String(256))
    reason = Column(String(256))
    site_domain = Column(String(256))
    owner = Column(String(256))
    company_name = Column(String(256))
    company_type = Column(String(256))
    main_page = Column(String(256))
    site_license = Column(String(256))
    site_name = Column(String(256))
    verify_time = Column(String(256))
    main_licence = Column(String(256))
    gmt_created = Column(String(256))
    gmt_updated = Column(String(256))


# 初始化数据库连接:
engine = create_engine('mysql+pymysql://root:BOOT-xwork1024@192.168.2.97:3306/spider')
# 创建DBSession类型:
DBSession = sessionmaker(bind=engine)
# 创建session对象:
session = DBSession()


def parse():
    # 获取用户目录
    LOCAL_FILE_PATH = input('输入文件')

    new_files = LOCAL_FILE_PATH.split('.')[0]

    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    workbook = open_workbook(LOCAL_FILE_PATH)  # 打开excel文件
    sheet2 = workbook.sheets()
    print(sheet2)
    # exit()
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)
    sum = []
    for i in range(2, sheet2.nrows):

        title_0 = sheet2.cell(i, 1).value
        title_1 = sheet2.cell(i, 2).value
        title_2 = sheet2.cell(i, 3).value
        title_3 = sheet2.cell(i, 4).value
        title_4 = sheet2.cell(i, 5).value
        title_5 = sheet2.cell(i, 6).value
        title_6 = sheet2.cell(i, 7).value
        title_7 = sheet2.cell(i, 8).value
        ti = title_1 + title_4 + title_7
        times = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        print(ti)
        if ti in sum:
            continue
        medi = Medicine(site_domain=title_0, company_name=title_1, company_type=title_2, owner=title_3
                        , site_license=title_4, site_name=title_5, main_page=title_6, verify_time=title_7,
                        gmt_created=times,gmt_updated=times)
        session.add(medi)
        sum.append(ti)
    session.commit()
    session.close()

def test():
    pass
if __name__ == '__main__':
    parse()

