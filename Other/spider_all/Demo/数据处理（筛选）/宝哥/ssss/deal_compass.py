import time, os
import logging.handlers
import openpyxl
from xlrd import open_workbook
import os

def parse():
    sum = []
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

    workbook = open_workbook("全部.xlsx")  # 打开excel文件
    sheet2 = workbook.sheets()
    print(sheet2)
    num = 1
    sheet2 = workbook.sheet_by_index(0)
    print(sheet2.nrows)

    for i in range(2, sheet2.nrows):
        # if i>1000:
        #     break
        title_0 = sheet2.cell(i, 12).value
        print(title_0)
        if title_0 == None or title_0 == "":
            continue

        sum = ['1',
               title_0,
               ]
        for index, s in enumerate(sum):
            outws.cell(num, index + 1).value = s  #
        num += 1

    outwb.save("a.xlsx")

def merge():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    from xlrd import open_workbook
    num = 1
    import os
    lis = os.listdir("D:\projects\S_Git_proj\spider\Other\spider_all\Demo\数据处理（筛选）\抖音数据\杭州")
    for l in lis:
        path = os.path.join('杭州',l)
        workbook = open_workbook(path)  # 打开excel文件
        sheet2 = workbook.sheets()
        print(sheet2)
        sheet2 = workbook.sheet_by_index(0)
        print(sheet2.nrows)
        for i in range(2, sheet2.nrows):
            title_0 = sheet2.cell(i, 9).value
            title_1 = sheet2.cell(i, 10).value
            title_2 = sheet2.cell(i, 18).value
            title_3 = sheet2.cell(i, 19).value
            print(title_0,title_1,title_2,title_3)


            sum = ['1',
                   title_0,
                   title_1,
                   title_2,
                   title_3

                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1
    outwb.save('{}_1.xlsx'.format('杭州'))

def excle_txt():
    lisss = os.listdir("D:\projects\S_Git_proj\spider\Other\spider_all\Demo\数据处理（筛选）\抖音数据\data")
    print(lisss)
    for a in lisss:
        # if a=='A12_filtered.xlsx':
        #     break
        workbook = open_workbook("data/{}".format(a))  # 打开excel文件
        sheet2 = workbook.sheets()
        print(sheet2)
        sheet2 = workbook.sheet_by_index(0)
        print(sheet2.nrows)
        for i in range(0, sheet2.nrows):

            title_0 = sheet2.cell(i, 1).value
            with open('data.txt','a',encoding='utf-8')as fp:
                fp.write(title_0+'\n')



import csv
def excle_csv():
    lisss = os.listdir("D:\projects\S_Git_proj\spider\Other\spider_all\Demo\数据处理（筛选）\抖音数据\data")
    print(lisss)
    for a in lisss:
        if a=='A12_filtered.xlsx':
            break
        workbook = open_workbook("data/{}".format(a))  # 打开excel文件
        sheet2 = workbook.sheets()
        print(sheet2)
        table = workbook.sheet_by_index(0)
        with open('data.csv', 'a', encoding='utf-8') as f:
            write = csv.writer(f)
            for row_num in range(table.nrows):
                row_value = table.row_values(row_num)
                write.writerow(row_value)



if __name__ == '__main__':
    parse()
    # excle_txt()