# coding=utf-8
import time, os
import logging.handlers
import openpyxl
# import pywintypes

####  pyinstaller 打包时出现问题，下载pywin32types即可，pip install pypiwin32

outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
from xlrd import open_workbook
workbook = open_workbook('../db_opera/dada.xlsx')  # 打开excel文件
all_sheet2 = workbook.sheets()
num = 1
sum1 = []

for ind,she in enumerate(all_sheet2):

    sheet2 = workbook.sheet_by_index(ind)  #
    rows = sheet2.nrows
    print(rows)

    for i in range(1, rows):
        count = 0
        title_0 = sheet2.cell(i, 1).value
        title_1 = sheet2.cell(i, 2).value # name
        title_2 = sheet2.cell(i, 3).value
        title_3 = sheet2.cell(i, 4).value
        title_4 = sheet2.cell(i, 5).value
        title_5 = sheet2.cell(i, 6).value
        title_6 = sheet2.cell(i, 7).value
        title_7 = sheet2.cell(i, 8).value

        for t in all_companys:
            if title_1 == t:
                count += 1
        if count > 1:
            # 公司名拿下来，查出来所在行数，对行数进行修改。
            continue

        if ind ==0:

            sum = ['1',
                   title_0,
                   title_1,
                   title_2,
                   title_3,
                   title_4,
                   title_5,
                   title_6,
                   title_7,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1
        if ind == 1:

            sum = ['1',None,title_1,None,None,None,None,None,None,
                   title_0,
                   title_2,
                   title_3,
                   title_4,
                   title_5,
                   title_6,
                   title_7,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1
        if ind == 2:

            sum = ['1',None,title_1,None,None,None,None,None,None,
                   None,None,None,None,None,None,None,
                   title_0,
                   title_2,
                   title_3,
                   title_4,
                   title_5,
                   title_6,
                   title_7,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1
        if ind == 3:

            sum = ['1',None,title_1,None,None,None,None,None,None
                ,None,None,None,None,None,None,None
                ,None,None,None,None,None,None,None,
                   title_0,
                   title_2,
                   title_3,
                   title_4,
                   title_5,
                   title_6,
                   title_7,
                   ]
            for index, s in enumerate(sum):
                outws.cell(num, index + 1).value = s  #
            num += 1

outwb.save('done.xlsx')
